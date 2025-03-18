import time
import os
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import urllib.parse
import re

def scraper_tokped(url,pages):
    # pages = int(input("berapa halaman yang ingin anda scrape (ex : 10) ? : "))
    print(f"🔍 Scraping: {url}")
    # Konfigurasi Selenium
    options = webdriver.ChromeOptions()
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

    driver = webdriver.Chrome(options=options)

    # Buka halaman Tokopedia
    # url = "https://www.tokopedia.com/search?fcity=174%2C175%2C176%2C177%2C178%2C179&navsource=&ob=9&sc=3058&srp_component_id=04.06.00.00&srp_page_id=&srp_page_title=&st=&q=samsung"
    driver.get(url)

    data = []

    # Loop untuk beberapa halaman (misal: 2 halaman)
    for i in range(pages):
        try:
            # Tunggu elemen produk muncul
            try:
                WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".css-rjanld")))
            except TimeoutException:
                print("❌ Error: Elemen tidak ditemukan atau halaman tidak dimuat dengan baik.")
                driver.quit()
                exit()

            # Scroll agar semua elemen muncul
            for j in range(10):  
                driver.execute_script("window.scrollBy(0, 500);")
                time.sleep(1)

            # Ambil HTML setelah scroll
            soup = BeautifulSoup(driver.page_source, 'html.parser')

            # Loop setiap produk
            for item in soup.find_all('div', class_='css-5wh65g'):
                try:
                    judul_elem = item.select_one('[class="_6+OpBPVGAgqnmycna+bWIw=="]').get_text(strip=True)
                    harga = item.select_one('[class="XvaCkHiisn2EZFq0THwVug=="]').get_text(strip=True)
                    # harga tanpa harga diskon
                    # harga = item.select_one('[class="_67d6E1xDKIzw+i2D2L0tjw=="]').get_text(strip=True)
                    url_toko = item.find('a')['href']
                    # Ambil informasi toko dan lokasi
                    for toko in item.select('[class="Jh7geoVa-F3B5Hk8ORh2qw=="]'):
                        nama_toko = toko.find('span').get_text(strip=True)
                        lokasi = toko.find_all('span')[1].get_text(strip=True)

                    terjual = "Tidak Ada terjual"  # Default sebelum iterasi
                    cust_rate_toko = "Tidak Ada rate"  # Default sebelum iterasi
                    # Ambil rating dan jumlah terjual
                    for rate in item.select('[class="Lrp+JcoWPuzTgMQ41Mkg3w=="]'):
                        if rate:  # Pastikan rate ditemukan
                            # Inisialisasi ulang nilai terjual di setiap iterasi
                            elemen_terjual = rate.select_one('[class="se8WAnkjbVXZNA8mT+Veuw=="]')
                            if elemen_terjual:
                                # print(f"Element ditemukan: {elemen_terjual}")
                                # print(f"Teks elemen: '{elemen_terjual.text.strip()}'")
                                terjual = elemen_terjual.get_text(strip=True) or "Tidak Ada"
                            else:
                                print("Element terjual tidak ditemukan")
                                terjual = "Tidak Ada"

                            # Ambil rating toko jika ada
                            rate_toko = rate.select_one('[class="_9jWGz3C-GX7Myq-32zWG9w=="]')
                            cust_rate_toko = rate_toko.get_text(strip=True) if rate_toko else "Tidak Adass"
                    product = {
                            'judul_elem': judul_elem,
                            'harga': harga,
                            'nama_toko': nama_toko,
                            'lokasi': lokasi,
                            'terjual': terjual,
                            'cust_rate_toko': cust_rate_toko,
                            'url': url_toko
                        }
                    data.append(product)
                    # df = pd.DataFrame(data,columns=('judul','harga','nama toko','lokasi','elemen_terjual','cust_rate_toko','url'))
                    # print(df)
                except AttributeError:
                    print("⚠️ Produk dengan struktur berbeda ditemukan, dilewati")

            # Coba klik tombol "Laman berikutnya"
            try:
                button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[aria-label="Laman berikutnya"]')))
                driver.execute_script("arguments[0].click();", button)
                time.sleep(3)
            except:
                print("⚠️ Tidak ada tombol 'Laman berikutnya'. Mungkin sudah halaman terakhir.")
                break

        except Exception as e:
            print(f"❌ Error di halaman {i+1}: {e}")
            break
    # Tutup browser
    driver.quit()
    return data
def editing_excel(filename):
                # Buka file Excel yang telah dibuat untuk editing
            wb = load_workbook(filename)
            ws = wb.active  # Ambil sheet aktif

            # **1. Auto-adjust lebar kolom berdasarkan panjang teks terpanjang**
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)  # Ambil nama kolom (A, B, C, ...)
                
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))  # Hitung panjang teks terpanjang
                    except:
                        pass
                
                ws.column_dimensions[col_letter].width = max_length + 2  # Tambahkan padding

            # **2. Atur alignment agar teks lebih rapi**
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            # **3. Terapkan Format Currency (Rp) pada kolom "harga"**
            for col in ws.iter_cols():
                header = col[0].value  # Ambil header kolom
                if header and header.lower() == "harga":  # Kolom harga
                    for cell in col[1:]:  # Lewati header
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '"Rp" #,##0'

            # Simpan perubahan ke file
            wb.save(filename)
def format_files(file):
    # Ambil query pencarian dari URL
    parsed_url = urllib.parse.urlparse(url)
    query_params = urllib.parse.parse_qs(parsed_url.query)
    search_keyword = query_params.get("q", ["output_data"])[0]  # Default ke "output_data" jika tidak ada query

    if file == "excel":
        print(f"EXCEL {file}")
        # Simpan ke Excel
        df.index = df.index + 1

        # Konversi kolom harga ke angka
        if "harga" in df.columns:
            df["harga"] = df["harga"].apply(lambda x: int(re.sub(r'[^\d]', '', x)) if isinstance(x, str) else x)

        # Cek apakah file sudah ada, jika iya, buat versi baru
        base_filename = search_keyword
        extension = ".xlsx"
        counter = 1
        filename = f"{base_filename}{extension}"

        while os.path.exists(filename):
            filename = f"{base_filename}({counter}){extension}"
            counter += 1

        df.to_excel(filename, index=False)
        editing_excel(filename)

    elif file == "csv":
        print(f"generate file {file} harap tunggu!!")
        # Simpan ke Excel
        df.index = df.index + 1

        # Cek apakah file sudah ada, jika iya, buat versi baru
        base_filename = search_keyword
        extension = ".csv"
        counter = 1
        filename = f"{base_filename}{extension}"

        while os.path.exists(filename):
            filename = f"{base_filename}({counter}){extension}"
            counter += 1
        # sep=";" untukagar ketika dibuka di excel kolom langsung rapi
        df.to_csv(filename,sep=";", index=False, encoding="utf-8")
    elif file == "json":
        print(f"generate file {file} harap tunggu!!")
        # Simpan ke json
        df.index = df.index + 1

        # Cek apakah file sudah ada, jika iya, buat versi baru
        base_filename = search_keyword
        extension = ".json"
        counter = 1
        filename = f"{base_filename}{extension}"

        while os.path.exists(filename):
            filename = f"{base_filename}({counter}){extension}"
            counter += 1
        # force_ascii=False hanya mencegah karakter ASCII-escaping, seperti: \/ → / (dalam URL)
        df.to_json(filename, orient="records", indent=4, force_ascii=False)
    os.system(f"start {filename}")
if __name__ == "__main__":
    pages = int(input("berapa halaman yang ingin anda scrape (ex : 10) ? : "))
    jumlah_url = int(input(f"Berapa banyak URL yang ingin di-scrape (gunakan angka): ")) 
    urls = []
    files = []
    for i in range(jumlah_url):
        url = input(f"\nMasukkan URL Tokopedia yang ingin di-scrape {i+1}: ")
        # scraper_tokped(url,pages)
        if "www.tokopedia.com" in url:
            urls.append(url)
        else :
            print(f"incorrect link tokopedia {i+1} skip link {i+1}")
            break
        format_file = input(f"Pilih Format File (CSV,EXCEL,JSON){i+1}: ").strip().lower()
        if format_file in ["csv", "excel", "json"] :
            files.append(format_file.strip())
        else : 
            print(f'masukan format file yang sesuai!! (CSV,EXCEL,JSON)')
            break
        # reset all data
        all_data = None
    for url,file in zip(urls,files):
        all_data = []
        url = url.strip()
            # print(f"🔍 Scraping: {url}")
        hasil = scraper_tokped(url,pages)
        all_data.extend(hasil)
        if all_data:
            df = pd.DataFrame(all_data)
            # all_data = None
        # print(df)
        # df.to_csv("hasil_scraping.csv", index=False)  # Simpan ke CSV
        else:
            print("❌ Tidak ada data yang berhasil diambil.")

        # extract to file csv,json, or excel
        # file = file.strip()
        print(f"format file {file} status : succes")
        format_files(file)


