import time
import os
import pandas as pd
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import urllib.parse
import shutil
import sqlalchemy
from dotenv import load_dotenv

def scraper_tokped(url,pages):

    if shutil.which("Xvfb"):
        os.system("killall -9 Xvfb 2>/dev/null || true")    # Hentikan Xvfb jika masih berjalan
        os.system("rm -f /tmp/.X99-lock")  # Hapus lock file
        os.system("Xvfb :99 -screen 0 1920x1080x24 &")
        os.environ["DISPLAY"] = ":99"

    # Konfigurasi Selenium
    options = webdriver.ChromeOptions()
    options.add_argument("user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")

    # ✅ Tambahkan opsi agar Chrome bisa jalan di Docker
    options.add_argument("--headless")  # Mode headless (tanpa UI)
    options.add_argument("--no-sandbox")  # Menghindari masalah sandbox di Docker
    options.add_argument("--disable-dev-shm-usage")  # Gunakan /tmp untuk shared memory
    options.add_argument("window-size=1920,1080")

    driver = webdriver.Chrome(options=options)

    # Buka halaman Tokopedia
    # url = "https://www.tokopedia.com/search?fcity=174%2C175%2C176%2C177%2C178%2C179&navsource=&ob=9&sc=3058&srp_component_id=04.06.00.00&srp_page_id=&srp_page_title=&st=&q=samsung"
    driver.get(url)

    data = []

    # Loop untuk beberapa halaman (misal: 2 halaman)
    for i in range(pages):
        try:
            # Tunggu elemen produk muncul
            try:
                WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CSS_SELECTOR, ".css-rjanld")))
            except TimeoutException:
                print("❌ Error: Elemen tidak ditemukan atau halaman tidak dimuat dengan baik.")
                driver.quit()
                exit()

            # Scroll agar semua elemen muncul
            for j in range(10):  
                driver.execute_script("window.scrollBy(0, 500);")
                time.sleep(1)

            # Ambil HTML setelah scroll
            soup = BeautifulSoup(driver.page_source, 'html.parser')

            # Loop setiap produk
            for item in soup.find_all('div', class_='css-5wh65g'):
                try:
                    items_title = item.select_one('[class="_6+OpBPVGAgqnmycna+bWIw=="]').get_text(strip=True)
                    url_store = item.find('a')['href']
                    # harga tanpa harga diskon
                    # harga = item.select_one('[class="_67d6E1xDKIzw+i2D2L0tjw=="]').get_text(strip=True)
                    # Ambil informasi toko dan lokasi
                    for pricelist in item.select('[class="XvaCkHiisn2EZFq0THwVug=="]'):
                        if pricelist :
                            normal_prices = pricelist.select('[class="hmtRf8oxRSR+n9OH5UxGoQ=="]')
                            # print("Discount elements found:", discounts)  # Debugging
                            normal_price = "tidak ada"
                            for price_normal in normal_prices:
                                normal_price =  price_normal.select_one('span')
                                if normal_price:
                                    normal_price = normal_price.get_text(strip=True) 
                                    # print("result normal: ",normal_price)
                                else:
                                    normal_price = "tidak ada"

                            discount_price_element = pricelist.select_one("div")
                            if discount_price_element:
                                discount_price = discount_price_element.get_text(strip=True) 
                            else :
                                discount_price_element = pricelist.select_one('[class="_67d6E1xDKIzw+i2D2L0tjw=="]')
                                if discount_price_element:
                                    discount_price = discount_price_element.get_text(strip=True)
                                else:
                                    discount_price = "Tidak ada harga diskon"

                    for store in item.select('[class="Jh7geoVa-F3B5Hk8ORh2qw=="]'):
                        store_name = store.find('span').get_text(strip=True)
                        location = store.find_all('span')[1].get_text(strip=True)

                    items_sold = ""  # Default sebelum iterasi
                    cust_rate = ""  # Default sebelum iterasi
                    # Ambil rating dan jumlah terjual
                    for rate in item.select('[class="Lrp+JcoWPuzTgMQ41Mkg3w=="]'):
                        if rate:  # Pastikan rate ditemukan
                            # Inisialisasi ulang nilai terjual di setiap iterasi
                            elements_sold = rate.select_one('[class="se8WAnkjbVXZNA8mT+Veuw=="]')
                            if elements_sold:
                                items_sold = elements_sold.get_text(strip=True) or "Tidak Ada"
                            else:
                                print("Element terjual tidak ditemukan")
                                items_sold = "Tidak Ada"

                            # Ambil rating toko jika ada
                            rate_store = rate.select_one('[class="_9jWGz3C-GX7Myq-32zWG9w=="]')
                            cust_rate = rate_store.get_text(strip=True) if rate_store else ""
                    product = {
                            'judul barang': items_title,
                            'harga diskon': discount_price,
                            'harga normal': normal_price,
                            'nama toko': store_name,
                            'lokasi': location,
                            'terjual': items_sold,
                            'custtomer rate': cust_rate,
                            'url barang': url_store
                        }
                    data.append(product)
                except AttributeError:
                    print("⚠️ Produk dengan struktur berbeda ditemukan, dilewati")

            # Coba klik tombol "Laman berikutnya"
            try:
                button = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'button[aria-label="Laman berikutnya"]')))
                driver.execute_script("arguments[0].click();", button)
                time.sleep(3)
            except:
                print("⚠️ Tidak ada tombol 'Laman berikutnya'. Mungkin sudah halaman terakhir.")
                break

        except Exception as e:
            print(f"❌ Error di halaman {i+1}: {e}")
            break
    # Tutup browser
    driver.quit()
    return data
def editing_excel(filename):
                # Buka file Excel yang telah dibuat untuk editing
            wb = load_workbook(filename)
            ws = wb.active  # Ambil sheet aktif

            # **1. Auto-adjust lebar kolom berdasarkan panjang teks terpanjang**
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)  # Ambil nama kolom (A, B, C, ...)
                
                for cell in col:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))  # Hitung panjang teks terpanjang
                    except:
                        pass
                
                ws.column_dimensions[col_letter].width = max_length + 2  # Tambahkan padding

            # **2. Atur alignment agar teks lebih rapi**
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='left', vertical='center')

            # **3. Terapkan Format Currency (Rp) pada kolom "harga"**
            for col in ws.iter_cols():
                header = col[0].value  # Ambil header kolom
                if header and header.lower() in  ["harga diskon", "harga normal"]:  # Kolom harga
                    for cell in col[1:]:  # Lewati header
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '"Rp" #,##0'

            # Simpan perubahan ke file
            wb.save(filename)
def get_unique_filename(filename):
    """Membuat nama file unik jika sudah ada"""
    counter = 1
    base, ext = os.path.splitext(filename)
    while os.path.exists(filename):
        filename = f"{base}({counter}){ext}"
        counter += 1
    return filename
def format_files(file, df, url):
    """Format output file sesuai tipe (Excel/JSON) dan pastikan nama unik"""
    parsed_url = urllib.parse.urlparse(url)
    query_params = urllib.parse.parse_qs(parsed_url.query)
    search_keyword = query_params.get("q", ["output_data"])[0]  # Default "output_data" jika tidak ada query
    base_filename = search_keyword

    # Pastikan extension sesuai jenis file
    extensions = {"excel": ".xlsx", "json": ".json","csv": ".csv","sql":"sql"}
    extension = extensions.get(file, ".txt")  # Default ke .txt jika tidak dikenal

    filename = f"output/{base_filename}{extension}"
    filename = get_unique_filename(filename)  # Pastikan nama file unik

    if file == "excel":
        df.to_excel(filename, index=False)
        editing_excel(filename)  # Fungsi tambahan
    elif file == "csv":
        df.to_csv(filename,sep=";", index=False, encoding="utf-8")
    elif file == "json":
        # force_ascii=False # hanya mencegah karakter ASCII-escaping, seperti: \/ → / (dalam URL)
        df.to_json(filename, orient="records", indent=4,force_ascii=False)
    elif file == "sql":
        load_dotenv() 
        password = os.getenv("PASSWORD")
        conn = sqlalchemy.create_engine(
            'mysql+mysqlconnector://root@localhost:3306/tokopedia')
            # f'mysql+mysqlconnector://ardi:{password}@localhost:3306/tokopedia')
        # force_ascii=False # hanya mencegah karakter ASCII-escaping, seperti: \/ → / (dalam URL)
        # Write the DataFrame to the database
        df.to_sql(name=search_keyword, con=conn, if_exists='replace', index=False)

        # Close the connection
        conn.dispose()
    print(f"✅ File berhasil dibuat: {filename}")
    return filename
if __name__ == "__main__":
    pages = int(input("berapa halaman yang ingin anda scrape? (gunakan angka) : "))
    jumlah_url = int(input(f"Berapa banyak URL yang ingin anda scrape? (gunakan angka): ")) 
    urls = []
    files = []
    for i in range(jumlah_url):
        url = input(f"\nMasukkan URL Tokopedia yang ingin di-scrape {i+1}: ")
        if "www.tokopedia.com" in url:
            urls.append(url)
        else :
            print(f"incorrect link tokopedia {i+1} skip link {i+1}")
            break
        format_file = input(f"Pilih Format File {i+1} (CSV,EXCEL,JSON,SQL) : ").strip().lower()
        if format_file in ["csv", "excel", "json","sql"] :
            files.append(format_file.strip())
        else : 
            print(f'masukan format file yang sesuai!! (CSV,EXCEL,JSON)')
            break
    
    # reset all data
    all_data = None
    for i, (url,file) in enumerate(zip(urls,files), start=1):
        print(f"\n🔍 Scraping link {i}")        
        all_data = []
        url = url.strip()
        hasil = scraper_tokped(url,pages)
        all_data.extend(hasil)
        if all_data:
            df = pd.DataFrame(all_data)
            # all_data = None
        # print(df)
        # df.to_csv("hasil_scraping.csv", index=False)  # Simpan ke CSV
        else:
            print("❌ Tidak ada data yang berhasil diambil.")

        # extract to file csv,json, or excel
        format_files(file,df,url)
        print(f"format file : {file} \nstatus : success\n")

